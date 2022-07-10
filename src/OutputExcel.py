import openpyxl as opx
import pprint as pp
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import shutil as sht
import datetime as dt
from copy import deepcopy

import Variables_List as VL 
import Pull_tabs_data as Ptd

class OutputWindow(tk.Frame):
    def __init__(self, master):
        super().__init__(master)

        self.names = Ptd.gettabsname()

        master.title("出力")

        MainFrame = tk.Frame(master)

        DirectoryFrame = tk.LabelFrame(MainFrame, text="出力先")
        DirectoryFrame2 = tk.Frame(DirectoryFrame)
        self.DirectoryEntry = tk.Entry(DirectoryFrame2, width=50)
        DEScrollbar = tk.Scrollbar(DirectoryFrame2, orient=tk.HORIZONTAL, command=self.DirectoryEntry.xview)
        self.DirectoryEntry["xscrollcommand"] = DEScrollbar.set
        DirectoryButton = ttk.Button(DirectoryFrame, text="参照", command=self.makeoutputfile)

        ComboboxFrame = tk.LabelFrame(MainFrame, text="出力プロジェクト")
        self.OutputCombobox = ttk.Combobox(ComboboxFrame, width=50, height=1, values=self.names)
        #RefreshButton = ttk.Button(ComboboxFrame, text="再取得", command=self.refreshnames)

        OutputButton = ttk.Button(MainFrame, text="出力", command=self.Output)

        self.DirectoryEntry.pack(side=tk.TOP)
        DEScrollbar.pack(side=tk.TOP, fill=tk.X)
        DirectoryFrame2.pack(side=tk.LEFT, anchor=tk.N)
        DirectoryButton.pack(side=tk.LEFT, anchor=tk.N)
        DirectoryFrame.pack(side=tk.TOP)

        self.OutputCombobox.pack(side=tk.LEFT)
        #RefreshButton.pack(side=tk.LEFT)
        ComboboxFrame.pack(side=tk.TOP, fill=tk.X)

        OutputButton.pack(side=tk.TOP, anchor=tk.E, fill=tk.X)

        MainFrame.pack()

    def makeoutputfile(self):
        OutputDirectory = filedialog.askdirectory(parent=self, title="出力先")
        self.DirectoryEntry.delete(0, tk.END)
        self.DirectoryEntry.insert(tk.END, str(OutputDirectory))

    def Output(self):
        OutputDirectory = self.DirectoryEntry.get()
        TabName = self.OutputCombobox.get()

        if TabName != "" and OutputDirectory != "":
            print("a")
            dt_now = dt.datetime.now()
            dt_str = str(dt_now.year) + str(dt_now.month) + str(dt_now.day) + str(dt_now.hour) + str(dt_now.minute)
            OutputDirectory2 = OutputDirectory + "/" + TabName + "(" + dt_str + ")" + ".xlsx"

            Gridsize = Ptd.pulltabsdata2(TabName)[2]
            Griddata = Ptd.pulltabsdata2(TabName)[3]
            Worddata = Ptd.pulltabsdata2(TabName)[7]
            
            #Copy xlsx files
            if Gridsize == 7:
                sht.copy("xlsx/7x7.xlsx", OutputDirectory2)
                output_word_list = deepcopy(VL.output_word_list_7)
            elif Gridsize == 12:
                sht.copy("xlsx/12x12.xlsx", OutputDirectory2)
                output_word_list = deepcopy(VL.output_word_list_12)
            elif Gridsize == 13:
                sht.copy("xlsx/13x13.xlsx", OutputDirectory2)
                output_word_list = deepcopy(VL.output_word_list_13)
            elif Gridsize == 20:
                sht.copy("xlsx/20x20.xlsx", OutputDirectory2)
                output_word_list = deepcopy(VL.output_word_list_20)

            #Load xlsx file
            wb = opx.load_workbook(OutputDirectory2)

            #Get sheets
            ws2 = wb.worksheets[0]
            ws1 = wb.worksheets[1]

            #Grid plot
            cou1 = 1
            for GriddataLine in Griddata:
                cou2 = 1
                for Grid in GriddataLine:
                    if Grid != 0:
                        ws1.cell(row=cou1, column=cou2, value=VL.output_word_id_list[int(Grid)-1])

                    if Grid == 1:
                        ws2.cell(row=cou1, column=cou2, value=VL.output_word_id_list[int(Grid)-1])
                    cou2 += 1
                cou1 += 1

            #Word list process
            for Wdata in Worddata:
                output_word_list[Wdata[1]-2].append(Wdata[0])

            pp.pprint(output_word_list)

            output_word_list_sorted = []
            for OWlist in output_word_list:
                if OWlist != []:
                    sorted_list = sorted(OWlist, reverse=True)
                    sorted_list.insert(0, "■" + str(len(OWlist[0])) + "文字")
                    output_word_list_sorted.append(sorted_list)

            pp.pprint(output_word_list_sorted)

            #Word list plot
            cou1 = 5
            for words in output_word_list_sorted:
                for word in words:
                    ws1.cell(row=Gridsize + cou1, column=1, value=word)
                    ws2.cell(row=Gridsize + cou1, column=1, value=word)
                    cou1 += 1

            #Save and close xlsx file
            wb.save(OutputDirectory2)
            wb.close()

        elif TabName == "" or OutputDirectory == "":
            tk.messagebox.showerror('注意', 'プロジェクトまたは出力先が選択されていません。')

def outputshow():
    window = tk.Toplevel()
    opwin = OutputWindow(master=window)
    opwin.mainloop()